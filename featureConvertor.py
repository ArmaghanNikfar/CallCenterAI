import openpyxl
from hijri_converter import convert
from datetime import datetime

# Mount Google Drive


def parse_date(cell_value):
    if isinstance(cell_value, datetime):
        return cell_value
    elif isinstance(cell_value, str):
        date_formats = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y")
        for fmt in date_formats:
            try:
                return datetime.strptime(cell_value, fmt)
            except ValueError:
                continue
    return None

def is_gregorian_leap_year(year):
    return (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)

def calculate_is_nowruz(gregorian_date_value):
    year = gregorian_date_value.year
    is_leap_year = is_gregorian_leap_year(year)
    if is_leap_year:
        nowruz_start = datetime(year, 3, 20)
        nowruz_end = datetime(year, 4, 1)
    else:
        nowruz_start = datetime(year, 3, 21)
        nowruz_end = datetime(year, 4, 2)
    return 1 if nowruz_start <= gregorian_date_value <= nowruz_end else 0

def calculate_is_summer(gregorian_date_value):
    year = gregorian_date_value.year
    is_leap_year = is_gregorian_leap_year(year)
    if is_leap_year:
        summer_start = datetime(year, 6, 21)
        summer_end = datetime(year, 9, 21)
    else:
        summer_start = datetime(year, 6, 22)
        summer_end = datetime(year, 9, 22)
    return 1 if summer_start <= gregorian_date_value <= summer_end else 0

def calculate_is_national_holiday(gregorian_date_value):
    year = gregorian_date_value.year
    is_leap_year = is_gregorian_leap_year(year)
    if is_leap_year:
        holidays = {
            datetime(year, 3, 20), datetime(year, 3, 21), datetime(year, 3, 22),
            datetime(year, 3, 23), datetime(year, 3, 31), datetime(year, 4, 1),
            datetime(year, 6, 3), datetime(year, 6, 4), datetime(year + 1, 2, 10),
            datetime(year + 1, 3, 20),
        }
    else:
        holidays = {
            datetime(year, 3, 21), datetime(year, 3, 22), datetime(year, 3, 23),
            datetime(year, 3, 24), datetime(year, 4, 1), datetime(year, 4, 2),
            datetime(year, 6, 4), datetime(year, 6, 5), datetime(year + 1, 2, 11),
            datetime(year + 1, 3, 19),
        }
    return 1 if gregorian_date_value in holidays else 0

def main():
    # Define input and output file paths
    input_file_path = '1402PureWithFeatures.xlsx'
    output_file_path = '1402To1403ChangeyearAndSetFeatures.xlsx'

    # Define specific Hijri dates (month and day) to be marked as religious holidays
    specific_hijri_dates = [

        (1, 9), (1, 10), (2, 20), (2, 28), (3, 8), (3, 17),
        (6, 3), (7, 13), (7, 27), (8, 15), (9, 21), (10, 1),
        (10, 2), (10, 25), (12, 10), (12, 18)
    ]

    try:
        # Load the input Excel file
        workbook = openpyxl.load_workbook(input_file_path)
        worksheet = workbook.active

        # Find the column indices by checking the header row
        column_indices = {
            "date": None,
            "isNowruz": None,
            "issummerholiday": None,
            "isramadan": None,
            "ismoharam": None,
            "isnationalholiday": None,
            "isreligiousholidays": None
        }

        for col_idx, cell in enumerate(worksheet[1], start=1):
            if cell.value:
                header = cell.value.strip().lower()
                if header in column_indices:
                    column_indices[header] = col_idx

        if column_indices["date"] is None:
            print("Date column not found in the input file.")
            return

        # Create new columns if they do not exist
        for key in column_indices:
            if key != "date" and column_indices[key] is None:
                column_indices[key] = worksheet.max_column + 1
                worksheet.cell(row=1, column=column_indices[key], value=key.capitalize())

        # Update the IsNowruz, IsSummerHoliday, IsRamadan, IsMoharam, IsNationalHoliday, and IsReligiousHolidays columns based on dates
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=column_indices["date"]).value
            cell_date = parse_date(cell_value)
            if cell_date:
                # Convert Gregorian to Hijri
                hijri_date = convert.Gregorian(cell_date.year, cell_date.month, cell_date.day).to_hijri()

                # Check if it's Ramadan (9th month of Hijri calendar)
                is_ramadan = hijri_date.month == 9

                # Check if it's Muharam (1st to 10th of the 1st month of Hijri calendar)
                is_moharam = hijri_date.month == 1 and 1 <= hijri_date.day <= 10

                # Check if it's a specific religious holiday
                is_religious_holiday = any(hijri_date.month == h_month and hijri_date.day == h_day for h_month, h_day in specific_hijri_dates)

                # Calculate Gregorian-based holidays
                is_nowruz = calculate_is_nowruz(cell_date)
                is_summer = calculate_is_summer(cell_date)
                is_national_holiday = calculate_is_national_holiday(cell_date)

                # Update the columns
                worksheet.cell(row=row_idx, column=column_indices["isNowruz"]).value = is_nowruz
                worksheet.cell(row=row_idx, column=column_indices["issummerholiday"]).value = is_summer
                worksheet.cell(row=row_idx, column=column_indices["isramadan"]).value = 1 if is_ramadan else 0
                worksheet.cell(row=row_idx, column=column_indices["ismoharam"]).value = 1 if is_moharam else 0
                worksheet.cell(row=row_idx, column=column_indices["isnationalholiday"]).value = is_national_holiday
                worksheet.cell(row=row_idx, column=column_indices["isreligiousholidays"]).value = 1 if is_religious_holiday else 0

        # Save the updated Excel file
        workbook.save(output_file_path)
        print("The Excel file has been successfully updated and saved as '{}'.".format(output_file_path))

    except FileNotFoundError:
        print("FileNotFoundError: The input file '{}' was not found.".format(input_file_path))
    except openpyxl.utils.exceptions.InvalidFileException:
        print("InvalidFileException: The input file '{}' is invalid.".format(input_file_path))
    except Exception as e:
        print("An error occurred: {}".format(e))

if __name__ == "__main__":
    main()
